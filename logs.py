import sys
import os
import dashboard
import pymysql
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
)
from PyQt5 import QtCore, QtGui, QtWidgets

from PyQt5.QtCore import QTimer
from datetime import datetime, timedelta

from PyQt5.QtWidgets import (
    QTableWidget,
    QScrollArea,
)
import openpyxl
from datetime import datetime

import login


class Logs(object):
    def __init__(self) -> None:
        self.logs_sent = False
        self.existing = False

        # For Excel File
        self.file_name = ""
        self.file_path = ""
        self.folder_name = "Logs"
        self.folder_path = rf"C:\Users\SampleUser\Desktop\{self.folder_name}"  # Change this to your own file path

        # Export and Delete Analytics Every Monday (Check Every Second)
        self.timer_second = 1
        self.timer = QTimer()
        self.timer.timeout.connect(self.analytics_reset_export_check)
        self.timer.start(self.timer_second * 1000)  # Execute every set self.time_second
        self.current_tbl_row_count = 0

    # Auto Refresh Table
    def analytics_reset_export_check(self):
        self.check_if_logs_has_new_value()
        print(self.current_tbl_row_count)

    def check_if_logs_has_new_value(self):
        connection = pymysql.connect(
            host="localhost", user="root", password="", db="suit_db"
        )
        cursor = connection.cursor()

        query = "SELECT count(id) FROM tbl_detect_log"
        cursor.execute(query)
        query_value = cursor.fetchone()[0]

        if query_value > self.current_tbl_row_count:
            self.search_logs()
        self.current_tbl_row_count = query_value
        cursor.close()

    def setupUi(self, MainWindow):
        self.MainWindow = MainWindow

        MainWindow.setObjectName("MainWindow")
        self.MainWindow.showMaximized()

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # Add table widget
        self.tableWidget = QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(310, 150, 1240, 900))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # Department Filter
        self.departments = ["ALL", "CABEIHM", "CAS", "CICS", "CET", "CONAHS", "CTE"]

        self.deparmentLabel = QtWidgets.QLabel(self.centralwidget)
        self.deparmentLabel.setGeometry(QtCore.QRect(600, 95, 70, 30))
        self.deparmentLabel.setObjectName("deparmentLabel")
        self.deparmentLabel.setText("Deparment:")

        self.deparmentComboBox = QtWidgets.QComboBox(self.centralwidget)
        self.deparmentComboBox.setGeometry(QtCore.QRect(670, 95, 100, 30))
        self.deparmentComboBox.setObjectName("deparmentComboBox")
        self.deparmentComboBox.addItems(self.departments)
        self.deparmentComboBox.currentTextChanged.connect(self.search_logs)

        # Detected Type Filter
        self.detected_type = ["ALL", "PROPER", "IMPROPER"]

        self.detectedtypeLabel = QtWidgets.QLabel(self.centralwidget)
        self.detectedtypeLabel.setGeometry(QtCore.QRect(800, 95, 70, 30))
        self.detectedtypeLabel.setObjectName("detectedtypeLabel")
        self.detectedtypeLabel.setText("Type:")

        self.detectedtypeComboBox = QtWidgets.QComboBox(self.centralwidget)
        self.detectedtypeComboBox.setGeometry(QtCore.QRect(840, 95, 100, 30))
        self.detectedtypeComboBox.setObjectName("detectedtypeComboBox")
        self.detectedtypeComboBox.addItems(self.detected_type)
        self.detectedtypeComboBox.currentTextChanged.connect(self.search_logs)

        # Date From Filter
        self.datefromLabel = QtWidgets.QLabel(self.centralwidget)
        self.datefromLabel.setGeometry(QtCore.QRect(960, 95, 70, 30))
        self.datefromLabel.setObjectName("datefromLabel")
        self.datefromLabel.setText("Date From:")

        self.datefromDateEdit = QtWidgets.QDateEdit(
            self.centralwidget, calendarPopup=True
        )
        self.datefromDateEdit.setGeometry(QtCore.QRect(1030, 95, 80, 30))
        self.datefromDateEdit.setObjectName("datefromDateEdit")
        self.datefromDateEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.datefromDateEdit.dateChanged.connect(self.search_logs)

        # Date To Filter
        self.datetoLabel = QtWidgets.QLabel(self.centralwidget)
        self.datetoLabel.setGeometry(QtCore.QRect(1130, 95, 70, 30))
        self.datetoLabel.setObjectName("datetoLabel")
        self.datetoLabel.setText("Date To:")

        self.datetoDateEdit = QtWidgets.QDateEdit(
            self.centralwidget, calendarPopup=True
        )
        self.datetoDateEdit.setGeometry(QtCore.QRect(1190, 95, 80, 30))
        self.datetoDateEdit.setObjectName("datetoDateEdit")
        self.datetoDateEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.datetoDateEdit.dateChanged.connect(self.search_logs)

        # search bar
        self.search_has_input = False

        self.searchLabel = QtWidgets.QLabel(self.centralwidget)
        self.searchLabel.setGeometry(QtCore.QRect(1310, 95, 70, 30))
        self.searchLabel.setObjectName("searchLabel")
        self.searchLabel.setText("Search:")

        self.searchLineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.searchLineEdit.setGeometry(QtCore.QRect(1360, 95, 150, 30))
        self.searchLineEdit.setObjectName("searchLineEdit")
        self.searchLineEdit.textChanged.connect(self.search_logs)

        self.exportDataBtn = QtWidgets.QPushButton(self.centralwidget)
        self.exportDataBtn.setGeometry(QtCore.QRect(1550, 90, 121, 40))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.exportDataBtn.setFont(font)
        self.exportDataBtn.setObjectName("exportDataBtn")
        self.exportDataBtn.setText("Export Data")
        self.exportDataBtn.setStyleSheet(
            """
            QPushButton {
                background-color: #dc3545;
                border: none;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-family: Arial;
                font-size: 8pt;
            }

            QPushButton:hover {
                background-color: #c82333;
            }
            """
        )
        self.exportDataBtn.clicked.connect(self.export_data_to_excel)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(-1, 1, 261, 2000))
        self.frame.setStyleSheet("background-color: #c82333;")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.label_4 = QtWidgets.QLabel(self.frame)
        self.label_4.setGeometry(QtCore.QRect(80, 20, 101, 91))
        self.label_4.setText("")
        self.label_4.setPixmap(QtGui.QPixmap("img/logo.png"))
        self.label_4.setScaledContents(True)
        self.label_4.setObjectName("label_4")
        self.label_9 = QtWidgets.QLabel(self.frame)
        self.label_9.setGeometry(QtCore.QRect(44, 110, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_9.setObjectName("label_9")

        self.dashboardBtn = QtWidgets.QPushButton(self.frame)
        self.dashboardBtn.setGeometry(QtCore.QRect(40, 200, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.dashboardBtn.setFont(font)
        self.dashboardBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.dashboardBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.dashboardBtn.setObjectName("dashboardBtn")

        self.logsBtn = QtWidgets.QPushButton(self.frame)
        self.logsBtn.setGeometry(QtCore.QRect(40, 250, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.logsBtn.setFont(font)
        self.logsBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.logsBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.logsBtn.setObjectName("logsBtn")

        self.detectionBtn = QtWidgets.QPushButton(self.frame)
        self.detectionBtn.setGeometry(QtCore.QRect(40, 300, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.detectionBtn.setFont(font)
        self.detectionBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.detectionBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.detectionBtn.setObjectName("detectionBtn")

        self.logoutBtn = QtWidgets.QPushButton(self.frame)
        self.logoutBtn.setGeometry(QtCore.QRect(40, 350, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.logoutBtn.setFont(font)
        self.logoutBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.logoutBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.logoutBtn.setObjectName("logoutBtn")
        self.logoutBtn.clicked.connect(self.open_login_page)

        self.frame_3 = QtWidgets.QFrame(self.centralwidget)
        self.frame_3.setGeometry(QtCore.QRect(261, -1, 2000, 61))
        self.frame_3.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.label = QtWidgets.QLabel(self.frame_3)
        self.label.setGeometry(QtCore.QRect(20, 10, 671, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        self.label_10.setGeometry(QtCore.QRect(310, 90, 270, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color:black;")
        self.label_10.setObjectName("label_10")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1186, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.detectionBtn.clicked.connect(self.open_detection)
        self.dashboardBtn.clicked.connect(self.open_dashboard)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(
            _translate(
                "MainWindow",
                "SUIT: School Uniform Identifier Technology using Object Detection",
            )
        )
        self.label_9.setText(_translate("MainWindow", "      SUIT"))
        self.dashboardBtn.setText(_translate("MainWindow", "Dashboard"))
        self.logsBtn.setText(_translate("MainWindow", "Improper Uniform Monitoring"))
        self.detectionBtn.setText(_translate("MainWindow", "Detection"))
        self.logoutBtn.setText(_translate("MainWindow", "Logout"))

        self.label.setText(
            _translate(
                "MainWindow",
                "SUIT: School Uniform Identifier Technology using Object Detection",
            )
        )
        self.label_10.setText(_translate("MainWindow", "Captured Improper Uniform"))

    def load_logs(self):
        if self.search_has_input:
            return

        print("Executed load_logs")

        connection = pymysql.connect(
            host="localhost", user="root", password="", db="suit_db"
        )
        cursor = connection.cursor()

        delete_query = "DELETE FROM tbl_logs WHERE date_log < %s"
        week_ago = datetime.now() - timedelta(days=7)
        cursor.execute(delete_query, (week_ago.date(),))
        connection.commit()

        query = "SELECT  unif_detect_result, department, course, date_log, time_log FROM tbl_detect_log ORDER BY ID DESC"
        cursor.execute(query)
        logs = cursor.fetchall()

        row_count = len(logs)
        column_count = 5  # Increase the column count for the image column

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)

        header_labels = [
            "Detected",
            "Department",
            "Course",
            "Date",
            "Time",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for row, log in enumerate(logs):
            unif_detect_reuslt, department, course, date_log, time_log = log

            self.tableWidget.setItem(row, 0, QTableWidgetItem(str(unif_detect_reuslt)))
            self.tableWidget.setItem(row, 1, QTableWidgetItem(str(department)))
            self.tableWidget.setItem(row, 2, QTableWidgetItem(str(course)))
            self.tableWidget.setItem(row, 3, QTableWidgetItem(str(date_log)))
            self.tableWidget.setItem(row, 4, QTableWidgetItem(str(time_log)))

        cursor.close()

    def search_logs(self):
        print("Department: ", self.deparmentComboBox.currentText())
        print("Type: ", self.detectedtypeComboBox.currentText())
        print("Date From: ", self.datefromDateEdit.date().toString("yyyy-MM-dd"))
        print("Date To: ", self.datefromDateEdit.date().toString("yyyy-MM-dd"))

        department_filter = self.deparmentComboBox.currentText()
        detectedtype_filter = self.detectedtypeComboBox.currentText()
        date_filter_from = self.datefromDateEdit.date().toString("yyyy-MM-dd")
        date_filter_to = self.datetoDateEdit.date().toString("yyyy-MM-dd")

        search_text = self.searchLineEdit.text()

        connection = pymysql.connect(
            host="localhost", user="root", password="", db="suit_db"
        )
        cursor = connection.cursor()

        if search_text:
            self.search_has_input = True
        else:
            self.search_has_input = False
            # QTimer.singleShot(1000, self.load_logs)

        query = """
        SELECT
           unif_detect_result, department, course, date_log,
           time_log
        FROM tbl_detect_log
        WHERE 1=1
        """
        data = []
        if department_filter != "ALL":
            query += " AND department = %s"
            data.append(department_filter)

        if detectedtype_filter != "ALL":
            query += " AND unif_detect_result = %s"
            data.append(detectedtype_filter)

        if date_filter_from and date_filter_to:
            query += " AND date_log BETWEEN %s AND %s"
            data.append(date_filter_from)
            data.append(date_filter_to)

        if search_text:
            query += " AND (course LIKE %s)"
            data.append(f"%{search_text}%")

        query += " ORDER BY ID DESC"

        print("Query", query)
        print("Data", data)

        cursor.execute(
            query,
            data,
        )
        logs = cursor.fetchall()

        row_count = len(logs)
        column_count = 5  # Increase the column count for the image column

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        header_labels = [
            "Detected",
            "Department",
            "Course",
            "Date",
            "Time",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for row, log in enumerate(logs):
            unif_detect_result, department, course, date_log, time_log = log

            self.tableWidget.setItem(row, 0, QTableWidgetItem(str(unif_detect_result)))
            self.tableWidget.setItem(row, 1, QTableWidgetItem(str(department)))
            self.tableWidget.setItem(row, 2, QTableWidgetItem(str(course)))
            self.tableWidget.setItem(row, 3, QTableWidgetItem(str(date_log)))
            self.tableWidget.setItem(row, 4, QTableWidgetItem(str(time_log)))

        cursor.close()
        connection.close()

    def start_loading_students(self):
        # Start loading the students initially
        self.search_logs()

    def open_detection(self):
        from UniformDetection import UniformDetectionWindow

        detection = UniformDetectionWindow()

        detection.uniform_detection_func()

    def open_dashboard(self):
        print("Opening Dashboard...")
        self.MainWindow.hide()
        self.dashboard_window = QtWidgets.QMainWindow()
        self.ui = dashboard.Ui_Dashboard()
        self.ui.setupUi(self.dashboard_window)
        self.dashboard_window.show()

    def open_login_page(self):
        print("Opening Login Page...")
        self.MainWindow.hide()
        self.login_window = QtWidgets.QMainWindow()
        self.ui = login.Ui_Form()
        self.ui.setupUi(self.login_window)
        self.login_window.show()

    def export_data_to_excel(self):
        # Connect to the MySQL database
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="suit_db"
        )
        department_filter = self.deparmentComboBox.currentText()
        detectedtype_filter = self.detectedtypeComboBox.currentText()
        date_filter_from = self.datefromDateEdit.date().toString("yyyy-MM-dd")
        date_filter_to = self.datetoDateEdit.date().toString("yyyy-MM-dd")

        search_text = self.searchLineEdit.text()

        try:
            # Create a cursor object to execute SQL queries
            cursor = connection.cursor()

            # Retrieve data from the tbl_student table
            query = """
                SELECT
                unif_detect_result, department, course, date_log,
                time_log
                FROM tbl_detect_log
                WHERE 1=1
            """
            data = []
            if department_filter != "ALL":
                query += " AND department = %s"
                data.append(department_filter)

            if detectedtype_filter != "ALL":
                query += " AND unif_detect_result = %s"
                data.append(detectedtype_filter)

            if date_filter_from and date_filter_to:
                query += " AND date_log BETWEEN %s AND %s"
                data.append(date_filter_from)
                data.append(date_filter_to)

            if search_text:
                query += " AND (course LIKE %s)"
                data.append(f"%{search_text}%")

            print("Query", query)
            print("Data", data)

            query += " ORDER BY ID DESC"
            cursor.execute(query, data)
            student_data = cursor.fetchall()

            # Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write the column headers
            sheet["A1"] = "Detected"
            sheet["B1"] = "Department"
            sheet["C1"] = "Course"
            sheet["D1"] = "Date Log"
            sheet["E1"] = "Time Log"

            # Set column width for date columns
            date_columns = ["D", "E"]  # Columns D and E represent the date columns
            for column in date_columns:
                sheet.column_dimensions[
                    column
                ].width = 15  # Adjust the width as per your preference

            for row_index, student in enumerate(student_data, start=2):
                sheet.cell(row=row_index, column=1).value = student[0]
                sheet.cell(row=row_index, column=2).value = student[1]
                sheet.cell(row=row_index, column=3).value = student[2]
                sheet.cell(row=row_index, column=4).value = student[3]
                sheet.cell(row=row_index, column=5).value = student[4]

            # Save the Excel file
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
            self.file_name = f"logs_data_{formatted_datetime}.xlsx"

            # Save the Excel file inside the "folder_path" folder
            self.file_path = rf"{self.folder_path}\{self.file_name}"

            # Create the "logs" folder if it doesn't exist
            if not os.path.exists(self.folder_path):
                os.makedirs(self.folder_path)
            # Save the Excel file
            workbook.save(self.file_path)
            print("Logs Data exported to Excel successfully!")

        except Exception as e:
            print("Error exporting data to Excel:", str(e))

        finally:
            # Close the cursor and connection
            cursor.close()
            connection.close()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Logs()
    ui.setupUi(MainWindow)
    ui.search_logs()
    MainWindow.show()
    sys.exit(app.exec_())
